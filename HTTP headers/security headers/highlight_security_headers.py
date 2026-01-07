import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ───────────────────────── FILES ─────────────────────────
CSV_FILE = "security_headers_report.csv"
EXCEL_FILE = "security_headers_report.xlsx"

# ───────────────────────── LOAD CSV ─────────────────────────
df = pd.read_csv(CSV_FILE)
df.to_excel(EXCEL_FILE, index=False)

# ───────────────────────── LOAD EXCEL ─────────────────────────
wb = load_workbook(EXCEL_FILE)
ws = wb.active

# ───────────────────────── COLORS ─────────────────────────
green_fill = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)  # Header present → GOOD

red_fill = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)  # Header missing → BAD

# ───────────────────────── HIGHLIGHT LOGIC ─────────────────────────
for row in ws.iter_rows(min_row=2, min_col=2):  # skip header row & domain column
    for cell in row:
        value = str(cell.value).strip().lower()

        if value == "oui":
            cell.fill = green_fill
        elif value == "non":
            cell.fill = red_fill

# ───────────────────────── SAVE ─────────────────────────
wb.save(EXCEL_FILE)
print(f"[+] Highlighted Excel report created: {EXCEL_FILE}")
